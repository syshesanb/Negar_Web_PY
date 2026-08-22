from typing import Generator
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session
from app.config import settings
from app.domain.models import Base, UserAccount, Company, FiscalYear, Permission, RolePermission
from app.infrastructure.security import hash_password

connect_args = {}
if settings.DATABASE_URL.startswith("sqlite"):
    connect_args = {"check_same_thread": False}

engine = create_engine(
    settings.DATABASE_URL,
    connect_args=connect_args,
    echo=False,
)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def get_db() -> Generator[Session, None, None]:
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def seed_default_chart_of_accounts(db: Session, company_id: int) -> int:
    """
    بارگذاری کدینگ استاندارد حسابداری از فایل cod_estandard.xlsx
    برای شرکت مشخص‌شده.
    سلسله‌مراتب: گروه → کل → معین
    در صورت وجود قبلی، بارگذاری مجدد انجام نمی‌شود.
    مقدار بازگشتی: تعداد حساب‌های بارگذاری‌شده
    """
    from app.domain.models import SarfaslHesab
    from pathlib import Path

    # بررسی وجود کدینگ برای این شرکت
    existing_count = db.query(SarfaslHesab).filter(SarfaslHesab.CompanyID == company_id).count()
    if existing_count > 0:
        return existing_count

    # مسیر فایل اکسل کدینگ استاندارد
    excel_path = Path(settings.BASE_DIR) / "cod_estandard.xlsx"
    if not excel_path.exists():
        return 0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active

        # نگاشت کدها به AccountID برای تعیین پدر
        code_to_id: dict = {}
        inserted = 0

        # مرحله ۱: جمع‌آوری گروه‌ها و کل‌های یکتا
        groups: dict = {}   # code -> name
        kols: dict = {}     # code -> (name, group_code)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            grp_code  = str(row[0]).strip() if row[0] else None
            grp_name  = str(row[1]).strip() if row[1] else None
            kol_code  = str(row[2]).strip() if row[2] else None
            kol_name  = str(row[3]).strip() if row[3] else None

            if grp_code and grp_name and grp_code not in groups:
                groups[grp_code] = grp_name
            if kol_code and kol_name and kol_code not in kols:
                kols[kol_code] = (kol_name, grp_code)

        # مرحله ۲: درج گروه‌ها (سطح گروه)
        for grp_code, grp_name in groups.items():
            existing = db.query(SarfaslHesab).filter(
                SarfaslHesab.CompanyID == company_id,
                SarfaslHesab.AccountCode == grp_code
            ).first()
            if existing:
                code_to_id[grp_code] = existing.AccountID
                continue
            acc = SarfaslHesab(
                CompanyID=company_id,
                AccountCode=grp_code,
                AccountName=grp_name,
                AccountType="گروه",
                ParentAccountID=None,
                IsActive=True,
                AccountNature="بدهکار/بستانکار",
            )
            db.add(acc)
            db.flush()
            code_to_id[grp_code] = acc.AccountID
            inserted += 1

        # مرحله ۳: درج کل‌ها (سطح کل)
        for kol_code, (kol_name, grp_code) in kols.items():
            existing = db.query(SarfaslHesab).filter(
                SarfaslHesab.CompanyID == company_id,
                SarfaslHesab.AccountCode == kol_code
            ).first()
            if existing:
                code_to_id[kol_code] = existing.AccountID
                continue
            parent_id = code_to_id.get(grp_code)
            acc = SarfaslHesab(
                CompanyID=company_id,
                AccountCode=kol_code,
                AccountName=kol_name,
                AccountType="کل",
                ParentAccountID=parent_id,
                IsActive=True,
                AccountNature="بدهکار/بستانکار",
            )
            db.add(acc)
            db.flush()
            code_to_id[kol_code] = acc.AccountID
            inserted += 1

        # مرحله ۴: درج معین‌ها (سطح معین)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            kol_code   = str(row[2]).strip() if row[2] else None
            moin_code  = str(row[4]).strip() if row[4] else None
            moin_name  = str(row[5]).strip() if row[5] else None
            nature     = str(row[6]).strip() if row[6] else "بدهکار/بستانکار"

            if not moin_code or not moin_name:
                continue

            existing = db.query(SarfaslHesab).filter(
                SarfaslHesab.CompanyID == company_id,
                SarfaslHesab.AccountCode == moin_code
            ).first()
            if existing:
                code_to_id[moin_code] = existing.AccountID
                continue

            parent_id = code_to_id.get(kol_code)
            acc = SarfaslHesab(
                CompanyID=company_id,
                AccountCode=moin_code,
                AccountName=moin_name,
                AccountType="معین",
                ParentAccountID=parent_id,
                IsActive=True,
                AccountNature=nature,
            )
            db.add(acc)
            db.flush()
            code_to_id[moin_code] = acc.AccountID
            inserted += 1

        db.commit()
        return inserted

    except Exception as e:
        db.rollback()
        print(f"[WARNING] خطا در بارگذاری کدینگ استاندارد: {e}")
        return 0


def init_db():
    """Create all tables and seed default admin user and company if not present."""
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    try:
        # Check if admin user exists
        admin = db.query(UserAccount).filter(UserAccount.Username == "admin").first()
        if not admin:
            hashed_pw = hash_password("admin123")
            admin_user = UserAccount(
                Username="admin",
                Password=hashed_pw,
                UserType="SuperAdmin",
                FullName="مدیر کل سیستم",
                IsActive=True,
                MaxCompaniesAllowed=99,
                MaxFiscalYearsPerCompany=99,
            )
            db.add(admin_user)
            db.commit()
            db.refresh(admin_user)

            # Create default company
            company = Company(
                CompanyName="شرکت نمونه نگار",
                CompanyCode="101",
                BrandName="نگار",
                EconomicCode="123456789012",
                IsActive=True,
                OwnerUserID=admin_user.UserID,
            )
            db.add(company)
            db.commit()
            db.refresh(company)

            # Create default fiscal year
            from datetime import datetime
            fiscal_year = FiscalYear(
                CompanyID=company.CompanyID,
                FiscalYearName="سال مالی ۱۴۰۵",
                StartDate=datetime(2026, 3, 21),
                EndDate=datetime(2027, 3, 20),
                IsActive=True,
            )
            db.add(fiscal_year)
            db.commit()

            # بارگذاری کدینگ استاندارد برای شرکت پیش‌فرض
            count = seed_default_chart_of_accounts(db, company.CompanyID)
            print(f"[INFO] کدینگ پیش‌فرض بارگذاری شد: {count} حساب برای شرکت {company.CompanyName}")

        else:
            # بررسی و بارگذاری کدینگ برای شرکت‌هایی که کدینگ ندارند
            from app.domain.models import SarfaslHesab
            companies = db.query(Company).filter(Company.IsActive == True).all()
            for comp in companies:
                acc_count = db.query(SarfaslHesab).filter(SarfaslHesab.CompanyID == comp.CompanyID).count()
                if acc_count == 0:
                    count = seed_default_chart_of_accounts(db, comp.CompanyID)
                    print(f"[INFO] کدینگ پیش‌فرض بارگذاری شد: {count} حساب برای شرکت {comp.CompanyName}")

        # Ensure new columns exist in Currencies table
        try:
            from sqlalchemy import text
            with engine.connect() as conn:
                for col, col_type in [
                    ("CbiRate", "NUMERIC(18, 4) DEFAULT 1.0"),
                    ("CbiRateDate", "VARCHAR(20)"),
                    ("TgjuRate", "NUMERIC(18, 4) DEFAULT 1.0"),
                    ("TgjuRateDate", "VARCHAR(20)"),
                    ("GlobalRate", "NUMERIC(18, 4) DEFAULT 1.0"),
                    ("GlobalRateDate", "VARCHAR(20)")
                ]:
                    try:
                        conn.execute(text(f"ALTER TABLE Currencies ADD COLUMN {col} {col_type};"))
                        conn.commit()
                    except Exception:
                        pass
        except Exception:
            pass

        # Seed Currencies if none exist
        from app.domain.models import Currency
        curr_count = db.query(Currency).count()
        if curr_count == 0:
            today_str = "1405/05/27"
            default_currencies = [
                Currency(CurrencyCode="IRR", CurrencyName="ریال ایران", CurrencySymbol="﷼", IsBase=True, ManualRate=1.0, ManualRateDate=today_str, CbiRate=1.0, TgjuRate=1.0, GlobalRate=1.0, OnlineRate=1.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="TMN", CurrencyName="تومان", CurrencySymbol="تومان", IsBase=False, ManualRate=10.0, ManualRateDate=today_str, CbiRate=10.0, TgjuRate=10.0, GlobalRate=10.0, OnlineRate=10.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="USD", CurrencyName="دلار آمریکا", CurrencySymbol="$", IsBase=False, ManualRate=615000.0, ManualRateDate=today_str, CbiRate=448500.0, TgjuRate=618500.0, GlobalRate=619200.0, OnlineRate=618500.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="EUR", CurrencyName="یورو", CurrencySymbol="€", IsBase=False, ManualRate=665000.0, ManualRateDate=today_str, CbiRate=486200.0, TgjuRate=668000.0, GlobalRate=669500.0, OnlineRate=668000.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="AED", CurrencyName="درهم امارات", CurrencySymbol="د.إ", IsBase=False, ManualRate=168000.0, ManualRateDate=today_str, CbiRate=122100.0, TgjuRate=168500.0, GlobalRate=168600.0, OnlineRate=168500.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="TRY", CurrencyName="لیر ترکیه", CurrencySymbol="₺", IsBase=False, ManualRate=18800.0, ManualRateDate=today_str, CbiRate=13750.0, TgjuRate=18950.0, GlobalRate=18920.0, OnlineRate=18950.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="CNY", CurrencyName="یوان چین", CurrencySymbol="¥", IsBase=False, ManualRate=86000.0, ManualRateDate=today_str, CbiRate=62100.0, TgjuRate=86700.0, GlobalRate=86800.0, OnlineRate=86700.0, OnlineRateDate=today_str),
                Currency(CurrencyCode="GBP", CurrencyName="پوند انگلیس", CurrencySymbol="£", IsBase=False, ManualRate=775000.0, ManualRateDate=today_str, CbiRate=569000.0, TgjuRate=776000.0, GlobalRate=778000.0, OnlineRate=776000.0, OnlineRateDate=today_str),
            ]
            db.add_all(default_currencies)
            db.commit()
    finally:
        db.close()
